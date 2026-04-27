# ============================================================
# export_manager.py — DocQA Case Engine v2.0
# Export test cases ke: Excel, PDF, CSV Jira, CSV TestRail
# ============================================================

import os
import sys
import csv
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import DB_PATH
from setup_db import get_connection

OUTPUT_DIR = "exports"


# ── Helpers ───────────────────────────────────────────────────

def fetch_scenarios() -> list[dict]:
    conn = get_connection()
    rows = conn.execute("""
        SELECT * FROM test_scenarios
        ORDER BY
            CASE risk_level
                WHEN 'Critical' THEN 1
                WHEN 'High'     THEN 2
                WHEN 'Medium'   THEN 3
                WHEN 'Low'      THEN 4
                ELSE 5
            END, risk_score DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M")


# ── 1. EXCEL EXPORT ───────────────────────────────────────────

def export_excel(scenarios: list[dict]) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: All Test Cases ──────────────────────────────
    ws = wb.active
    ws.title = "Test Cases"

    RISK_COLORS = {
        "Critical":   "EF4444",
        "High":       "F97316",
        "Medium":     "EAB308",
        "Low":        "22C55E",
        "Unassessed": "94A3B8",
    }
    STATUS_COLORS = {
        "Approved": "22C55E",
        "Pending":  "94A3B8",
        "Rejected": "EF4444",
    }

    # Header row
    headers = [
        "ID", "Feature", "Scenario Title", "Test Type",
        "Preconditions", "Test Steps", "Expected Result",
        "Risk Level", "Risk Score", "Probability", "Impact",
        "Risk Reasoning", "Status", "Source", "LLM Model", "Created At"
    ]

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="E2E8F0", name="Arial", size=10)
    thin = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 32

    # Data rows
    for row_idx, sc in enumerate(scenarios, 2):
        risk = sc.get("risk_level", "Unassessed")
        status = sc.get("status", "Pending")

        values = [
            sc.get("id"),
            sc.get("feature_name", ""),
            sc.get("scenario_title", ""),
            sc.get("test_type", ""),
            sc.get("preconditions", "") or "",
            sc.get("test_steps", "") or "",
            sc.get("expected_result", "") or "",
            risk,
            sc.get("risk_score", 0),
            sc.get("probability_of_failure", 0),
            sc.get("business_impact", 0),
            sc.get("risk_reasoning", "") or "",
            status,
            sc.get("source", ""),
            sc.get("llm_model", "") or "",
            sc.get("created_at", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            # Color risk level cell
            if col == 8 and risk in RISK_COLORS:
                cell.fill = PatternFill("solid", fgColor=RISK_COLORS[risk])
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)

            # Color status cell
            if col == 13 and status in STATUS_COLORS:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)

            # Zebra stripe
            if row_idx % 2 == 0 and col not in (8, 13):
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

        ws.row_dimensions[row_idx].height = 60

    # Column widths
    col_widths = [6, 18, 35, 14, 28, 40, 35, 12, 9, 11, 9, 35, 12, 14, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Risk Summary ────────────────────────────────
    ws2 = wb.create_sheet("Risk Summary")
    ws2["A1"] = "DocQA Case Engine v2.0 — Risk Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial", color="0F172A")
    ws2["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    ws2["A2"].font = Font(size=10, name="Arial", color="64748B")

    ws2["A4"] = "Risk Level"
    ws2["B4"] = "Count"
    ws2["C4"] = "Percentage"
    for col in ["A4", "B4", "C4"]:
        ws2[col].font = Font(bold=True, name="Arial", color="FFFFFF")
        ws2[col].fill = PatternFill("solid", fgColor="1E293B")
        ws2[col].alignment = Alignment(horizontal="center")

    risk_counts = {}
    for sc in scenarios:
        rl = sc.get("risk_level", "Unassessed")
        risk_counts[rl] = risk_counts.get(rl, 0) + 1

    total = len(scenarios)
    for r_idx, (level, color) in enumerate([
        ("Critical", "EF4444"), ("High", "F97316"),
        ("Medium", "EAB308"), ("Low", "22C55E"), ("Unassessed", "94A3B8")
    ], 5):
        count = risk_counts.get(level, 0)
        ws2.cell(r_idx, 1, level).fill = PatternFill("solid", fgColor=color)
        ws2.cell(r_idx, 1).font = Font(bold=True, color="FFFFFF", name="Arial")
        ws2.cell(r_idx, 2, count).font = Font(name="Arial")
        ws2.cell(r_idx, 2).alignment = Alignment(horizontal="center")
        pct = f"=B{r_idx}/SUM(B5:B9)*100"
        ws2.cell(r_idx, 3, pct).number_format = "0.0%"
        ws2.cell(r_idx, 3).font = Font(name="Arial")
        ws2.cell(r_idx, 3).alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 14

    # Save
    path = os.path.join(OUTPUT_DIR, f"docqa_test_cases_{timestamp()}.xlsx")
    wb.save(path)
    return path


# ── 2. PDF EXPORT ─────────────────────────────────────────────

def export_pdf(scenarios: list[dict]) -> str:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    path = os.path.join(OUTPUT_DIR, f"docqa_test_report_{timestamp()}.pdf")

    RISK_COLORS_RL = {
        "Critical":   colors.HexColor("#EF4444"),
        "High":       colors.HexColor("#F97316"),
        "Medium":     colors.HexColor("#EAB308"),
        "Low":        colors.HexColor("#22C55E"),
        "Unassessed": colors.HexColor("#94A3B8"),
    }
    STATUS_COLORS_RL = {
        "Approved": colors.HexColor("#22C55E"),
        "Pending":  colors.HexColor("#94A3B8"),
        "Rejected": colors.HexColor("#EF4444"),
    }

    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=18, fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#0F172A"), spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   fontSize=9,  fontName="Helvetica",
                                  textColor=colors.HexColor("#64748B"), spaceAfter=12)
    cell_style  = ParagraphStyle("cell",  fontSize=7.5, fontName="Helvetica",
                                  textColor=colors.HexColor("#1E293B"), leading=11)
    bold_cell   = ParagraphStyle("bold",  fontSize=7.5, fontName="Helvetica-Bold",
                                  textColor=colors.white)

    story = []

    # Cover
    story.append(Paragraph("DocQA Case Engine v2.0", title_style))
    story.append(Paragraph("Test Case Report — Risk-Based Testing", sub_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  ·  Total: {len(scenarios)} test cases", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0"), spaceAfter=12))

    # Summary stats
    risk_counts = {}
    status_counts = {}
    for sc in scenarios:
        risk_counts[sc.get("risk_level","Unassessed")] = risk_counts.get(sc.get("risk_level","Unassessed"), 0) + 1
        status_counts[sc.get("status","Pending")] = status_counts.get(sc.get("status","Pending"), 0) + 1

    summary_data = [["Risk Level", "Count"]] + [
        [level, risk_counts.get(level, 0)]
        for level in ["Critical", "High", "Medium", "Low", "Unassessed"]
    ]
    summary_table = Table(summary_data, colWidths=[60*mm, 30*mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#1E293B")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("ALIGN",       (1,0), (1,-1), "CENTER"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#F8FAFC"), colors.white]),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 16))

    # Main table header
    story.append(Paragraph("All Test Scenarios", ParagraphStyle("h2", fontSize=11,
                            fontName="Helvetica-Bold", textColor=colors.HexColor("#1E293B"), spaceAfter=8)))

    col_headers = ["#", "Feature", "Scenario Title", "Type", "Preconditions",
                   "Test Steps", "Expected Result", "Risk", "Score", "Status"]
    col_widths_pdf = [8*mm, 25*mm, 48*mm, 18*mm, 38*mm, 52*mm, 45*mm, 16*mm, 10*mm, 16*mm]

    table_data = [[Paragraph(f"<b>{h}</b>", ParagraphStyle("th", fontSize=7.5,
                   fontName="Helvetica-Bold", textColor=colors.white)) for h in col_headers]]

    for sc in scenarios:
        risk  = sc.get("risk_level", "Unassessed")
        status = sc.get("status", "Pending")
        rc = RISK_COLORS_RL.get(risk, colors.HexColor("#94A3B8"))
        sc_txt = RISK_COLORS_RL.get(risk, colors.HexColor("#94A3B8"))

        def p(text): return Paragraph(str(text or ""), cell_style)

        table_data.append([
            p(sc.get("id","")),
            p(sc.get("feature_name","")),
            p(sc.get("scenario_title","")),
            p(sc.get("test_type","")),
            p(sc.get("preconditions","") or ""),
            p(sc.get("test_steps","") or ""),
            p(sc.get("expected_result","") or ""),
            Paragraph(f"<b>{risk}</b>", ParagraphStyle("risk", fontSize=7,
                fontName="Helvetica-Bold", textColor=colors.white,
                backColor=rc, borderPadding=2)),
            p(sc.get("risk_score",0)),
            Paragraph(f"<b>{status}</b>", ParagraphStyle("stat", fontSize=7,
                fontName="Helvetica-Bold", textColor=colors.white,
                backColor=STATUS_COLORS_RL.get(status, colors.HexColor("#94A3B8")),
                borderPadding=2)),
        ])

    main_table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    main_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#1E293B")),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.HexColor("#F8FAFC"), colors.white]),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#E2E8F0")),
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
    ]))

    story.append(main_table)

    doc.build(story)
    return path


# ── 3. CSV JIRA ───────────────────────────────────────────────
# Format: Summary, Issue Type, Priority, Description, Labels, Acceptance Criteria

def export_csv_jira(scenarios: list[dict]) -> str:
    path = os.path.join(OUTPUT_DIR, f"docqa_jira_{timestamp()}.csv")

    JIRA_PRIORITY = {
        "Critical": "Highest",
        "High":     "High",
        "Medium":   "Medium",
        "Low":      "Low",
        "Unassessed": "Low",
    }

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Summary",
            "Issue Type",
            "Priority",
            "Labels",
            "Description",
            "Acceptance Criteria",
            "Custom Field (Risk Level)",
            "Custom Field (Risk Score)",
            "Custom Field (Test Type)",
            "Custom Field (Feature)",
        ])

        for sc in scenarios:
            risk  = sc.get("risk_level", "Unassessed")
            steps = sc.get("test_steps", "") or ""
            pre   = sc.get("preconditions", "") or ""
            exp   = sc.get("expected_result", "") or ""
            reasoning = sc.get("risk_reasoning", "") or ""

            description = (
                f"*Feature:* {sc.get('feature_name','')}\n\n"
                f"*Preconditions:*\n{pre}\n\n"
                f"*Test Steps:*\n{steps}\n\n"
                f"*Risk Reasoning:* {reasoning}"
            )

            writer.writerow([
                f"[{sc.get('test_type','')}] {sc.get('scenario_title','')}",
                "Test",
                JIRA_PRIORITY.get(risk, "Medium"),
                f"qa,{risk.lower()},{sc.get('feature_name','').lower().replace(' ','-')}",
                description,
                exp,
                risk,
                sc.get("risk_score", 0),
                sc.get("test_type", ""),
                sc.get("feature_name", ""),
            ])

    return path


# ── 4. CSV TESTRAIL ───────────────────────────────────────────
# Format sesuai TestRail CSV import template

def export_csv_testrail(scenarios: list[dict]) -> str:
    path = os.path.join(OUTPUT_DIR, f"docqa_testrail_{timestamp()}.csv")

    TESTRAIL_PRIORITY = {
        "Critical": "1 - Critical",
        "High":     "2 - High",
        "Medium":   "3 - Medium",
        "Low":      "4 - Low",
        "Unassessed": "4 - Low",
    }
    TESTRAIL_TYPE = {
        "Positive":   "Functional",
        "Negative":   "Functional",
        "Boundary":   "Functional",
        "Edge Case":  "Functional",
    }

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        # TestRail standard import columns
        writer.writerow([
            "Title",
            "Section",
            "Type",
            "Priority",
            "Estimate",
            "References",
            "Preconditions",
            "Steps",
            "Expected Result",
            "Custom (Risk Level)",
            "Custom (Risk Score)",
            "Custom (Risk Reasoning)",
        ])

        for sc in scenarios:
            risk     = sc.get("risk_level", "Unassessed")
            test_type = sc.get("test_type", "Positive")
            steps    = sc.get("test_steps", "") or ""
            pre      = sc.get("preconditions", "") or ""
            exp      = sc.get("expected_result", "") or ""
            reasoning = sc.get("risk_reasoning", "") or ""

            # TestRail steps format: numbered list
            formatted_steps = steps  # already newline-separated from LLM

            writer.writerow([
                sc.get("scenario_title", ""),
                sc.get("feature_name", ""),         # Maps to Section/Suite
                TESTRAIL_TYPE.get(test_type, "Functional"),
                TESTRAIL_PRIORITY.get(risk, "3 - Medium"),
                "",                                  # Estimate (optional)
                "",                                  # References/Jira ticket
                pre,
                formatted_steps,
                exp,
                risk,
                sc.get("risk_score", 0),
                reasoning,
            ])

    return path


# ── CLI ───────────────────────────────────────────────────────

def run_export(formats: list[str] = None):
    ensure_output_dir()
    scenarios = fetch_scenarios()

    if not scenarios:
        print("⚠️  Tidak ada test case di database. Jalankan llm_intake.py dulu.")
        return

    print(f"\n📦 Exporting {len(scenarios)} test cases...\n")

    results = {}
    all_formats = formats or ["excel", "pdf", "jira", "testrail"]

    if "excel" in all_formats:
        print("  📊 Excel...", end=" ", flush=True)
        try:
            p = export_excel(scenarios)
            results["excel"] = p
            print(f"✅ {p}")
        except Exception as e:
            print(f"❌ {e}")

    if "pdf" in all_formats:
        print("  📄 PDF...", end=" ", flush=True)
        try:
            p = export_pdf(scenarios)
            results["pdf"] = p
            print(f"✅ {p}")
        except Exception as e:
            print(f"❌ {e}")

    if "jira" in all_formats:
        print("  🔵 CSV Jira...", end=" ", flush=True)
        try:
            p = export_csv_jira(scenarios)
            results["jira"] = p
            print(f"✅ {p}")
        except Exception as e:
            print(f"❌ {e}")

    if "testrail" in all_formats:
        print("  🟢 CSV TestRail...", end=" ", flush=True)
        try:
            p = export_csv_testrail(scenarios)
            results["testrail"] = p
            print(f"✅ {p}")
        except Exception as e:
            print(f"❌ {e}")

    print(f"\n✅ Export selesai! File tersimpan di folder: {os.path.abspath(OUTPUT_DIR)}/")
    return results


if __name__ == "__main__":
    print("\n" + "=" * 55)
    print("  DocQA Case Engine v2.0 — Export Manager")
    print("=" * 55)
    print("\nExport format:")
    print("  1. Semua (Excel + PDF + CSV Jira + CSV TestRail)")
    print("  2. Excel saja")
    print("  3. PDF saja")
    print("  4. CSV Jira saja")
    print("  5. CSV TestRail saja")
    print("  6. CSV Jira + TestRail saja")

    choice = input("\nPilihan (1-6): ").strip()

    fmt_map = {
        "1": ["excel", "pdf", "jira", "testrail"],
        "2": ["excel"],
        "3": ["pdf"],
        "4": ["jira"],
        "5": ["testrail"],
        "6": ["jira", "testrail"],
    }

    formats = fmt_map.get(choice)
    if not formats:
        print("❌ Pilihan tidak valid.")
        sys.exit(1)

    run_export(formats)